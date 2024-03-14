from pdf2docx import Converter
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import Workbook
from PyPDF2 import PdfReader

janela = tk.Tk()
janela.title("Conversor PDF para Docx e Xlsx")
largura_janela = 400
altura_janela = 180
pos_x = (janela.winfo_screenwidth() // 2) - (largura_janela // 2)
pos_y = (janela.winfo_screenheight() // 2) - (altura_janela // 2)
janela.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")

arquivo_selecionado = ""

def selecionar_arquivo():
    global arquivo_selecionado
    arquivo_selecionado = filedialog.askopenfilename()
    arquivo_nome = os.path.basename(arquivo_selecionado)
    label_arquivo.config(text="Arquivo selecionado: " + arquivo_nome)

def converter_arquivo():
    global arquivo_selecionado
    pdf_file = arquivo_selecionado
    docx_file = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Document", "*.docx")])
    if docx_file: 
        cv = Converter(pdf_file)
        cv.convert(docx_file)
        label_converter.config(text="Arquivo convertido com sucesso!")
        label_convertido.config(text="Deseja visualizar o arquivo convertido?")
        botao_convertido = tk.Button(janela, text="Visualizar Arquivo", command=lambda: visualizar_arquivo(docx_file))
        botao_convertido.place(x=240, y=140)
        sair = tk.Button(janela, text="  Sair  ", command=exit)
        sair.place(x=350, y=140)
        cv.close()

def converter_excel():
    global arquivo_selecionado
    pdf_file = arquivo_selecionado
    excel_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
    if excel_file:
        wb = Workbook()
        ws = wb.active
        ws.title = "Conteúdo do PDF"
        reader = PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        ws['A1'] = "Conteúdo do PDF:"
        ws['A2'] = text
        wb.save(excel_file)
        botao_convertido = tk.Button(janela, text="Visualizar Arquivo", command=lambda: visualizar_arquivo(excel_file))
        botao_convertido.place(x=240, y=140)
        sair = tk.Button(janela, text="  Sair  ", command=exit)
        sair.place(x=350, y=140)
        label_converter.config(text="Arquivo convertido com sucesso!")
        label_convertido.config(text="Deseja visualizar o arquivo convertido?")

def visualizar_arquivo(arquivo):
    os.system(f'start "" "{arquivo}"')


botao_selecionar = tk.Button(janela, text="Selecionar Arquivo", command=selecionar_arquivo)
botao_selecionar.place(x=20, y=20)

label_arquivo = tk.Label(janela, text="")
label_arquivo.place(x=20, y=50)

botao_converter = tk.Button(janela, text="Converter para Word", command=converter_arquivo)
botao_converter.place(x=20, y=80)

botao_excel = tk.Button(janela, text="Converter para Excel", command=converter_excel)
botao_excel.place(x=150, y=80)

label_converter = tk.Label(janela, text="")
label_converter.place(x=20, y=110)

label_convertido = tk.Label(janela, text="")
label_convertido.place(x=20, y=140)

janela.mainloop()

